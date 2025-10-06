import openpyxl
import re
from openpyxl.styles import Font, Border, Alignment, Side
from openpyxl.worksheet.worksheet import Worksheet
from typing import List, Any, Tuple, Dict
from pathlib import Path
from .plan_structure import ReleaseStructure, PhaseTeamComposition
from config.logger import setup_logger

logger = setup_logger("ExcelGenerator")

thin_border = Side(border_style="thin", color="000000")
border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)

def _create_regular_cell(row: int, column: int, ws: Worksheet, value: Any):
    cell = ws.cell(row=row, column=column, value=value)
    cell.border = border
    cell.font = Font(size=12)

def _create_column_name_cell(row: int, column: int, ws: Worksheet, value: str):
    cell = ws.cell(row=row, column=column, value=re.sub("_"," ", string=value).title())
    cell.border = border
    cell.font = Font(bold=True, size=12)

def _create_title_1(row: int, column: int, ws: Worksheet, title: str):
    title_cell = ws.cell(row=row, column=column, value=title.upper())
    title_cell.font = Font(bold=True,size=16)
    title_cell.border = border

def _create_title_2(row: int, column: int, ws: Worksheet, sub_title: str):
    title_cell = ws.cell(row=row, column=column, value=re.sub("_", " ", string=sub_title).title())
    title_cell.font = Font(bold=True,size=14)
    title_cell.border = border

def _create_title_3(row: int, column: int, ws: Worksheet, sub_title: str):
    title_cell = ws.cell(row=row, column=column, value=re.sub("_", " ", string=sub_title).title())
    title_cell.font = Font(bold=True,size=13)
    title_cell.border = border

def _write_column_names(ws: Worksheet, column_names: List[str], row: int):
    for i in range(len(column_names)):
        _create_column_name_cell(
            row=row, 
            column=i+1, 
            ws=ws, 
            value=column_names[i].upper())

def _construct_table(
        key: str, 
        ws: Worksheet, 
        release: Dict[str, List[Dict]],
        start_row: int) -> int:
    rows = release[key]
    col_names = list(rows[0].keys())
    _create_title_2(row=start_row, column=1, ws=ws, sub_title=key)
    start_row += 1
    _write_column_names(ws=ws, column_names=col_names, row=start_row)
    start_row += 1
    for row in rows:
        values = list(row.values())
        for i in range(len(col_names)):
            _create_regular_cell(
                row=start_row,
                column=i+1,
                ws=ws,
                value=values[i]
            )

        start_row += 1
    
    return start_row + 1

def _construct_phase_team_composition(
        ws: Worksheet,
        start_row: int,
        teams_list: List[PhaseTeamComposition]) -> int:
    for team in teams_list:
        _create_title_3(row=start_row, column=1, ws=ws, sub_title=team.phase)
        start_row += 1
        col_names = list(team.teams[0].model_dump().keys())
        _write_column_names(
            ws=ws,
            column_names= col_names,
            row=start_row
        )
        start_row += 1
        for row in team.teams:
            values = list(row.model_dump().values())
            for i in range(len(col_names)):
                _create_regular_cell(
                    row=start_row,
                    column=i+1,
                    ws=ws,
                    value=values[i]
                )
            start_row += 1
    
    return start_row

def _construct_geographic_distribution(
        roles: List[str],
        team_list: List[PhaseTeamComposition],
        start_row: int,
        ws: Worksheet,
        hours_per_worker: int = 80  # Assuming 80 hours per sprint per worker
) -> int:
    # Write header
    _create_title_2(row=start_row, column=1, ws=ws, sub_title="Geographic Distribution")
    start_row += 1
    
    # Write column names
    col_names = ['Role', 'Total Hours', 'Onshore Hours', 'Offshore Hours']
    _write_column_names(ws=ws, column_names=col_names, row=start_row)
    start_row += 1
    
    # Calculate hours for each role
    for role in roles:
        onshore_hours = 0
        offshore_hours = 0
        
        # Iterate through all phases and teams
        for phase_comp in team_list:
            for team in phase_comp.teams:
                # Get the headcount for this role in this team
                headcount = getattr(team, role, 0)
                
                # Accumulate hours based on team location
                if team.team == 'Onshore':
                    onshore_hours += headcount * hours_per_worker
                elif team.team == 'Offshore':
                    offshore_hours += headcount * hours_per_worker
        
        # Calculate total hours
        total_hours = onshore_hours + offshore_hours
        
        # Write row data
        _create_regular_cell(row=start_row, column=1, ws=ws, value=role)
        _create_regular_cell(row=start_row, column=2, ws=ws, value=total_hours)
        _create_regular_cell(row=start_row, column=3, ws=ws, value=onshore_hours)
        _create_regular_cell(row=start_row, column=4, ws=ws, value=offshore_hours)
        
        start_row += 1
    
    return start_row
            


def file_generator(releases: List[ReleaseStructure], output_path: str):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    # creating new named sheet
    ws = wb.create_sheet(title="ResourcePlan")
    logger.info("Starting file generation")
    current_row = 1
    #current_col = 1
    for release in releases:
        header = f"Release {release.release_number}"
        _create_title_1(row=current_row, column=1, ws=ws, title=header)
        logger.info(f"Writing for {header}")
        current_row += 1
        # Writing Phase Distribution tables and Sprint Allocation
        known_tables = ['phase_distribution', 'sprint_allocation']
        for key in known_tables:
            logger.info(f"Creating table for {key}")
            current_row = _construct_table(
                key=key,
                ws=ws,
                release=release.model_dump(),
                start_row=current_row
            )
            logger.info("Table succesfully created")
        current_row += 1
        # writing table Team Composition
        _create_title_2(row=current_row, column=1, ws=ws, sub_title="Team Composition (by Phase)")
        current_row += 1
        current_row = _construct_phase_team_composition(ws=ws, start_row=current_row, teams_list=release.team_composition) 
        logger.info("Table created for team composition")
        current_row += 2

        current_row = _construct_table(
            key="geographic_distribution",
            ws=ws,
            release=release.model_dump(),
            start_row=current_row
        )
        logger.info("Table Geographic Distribution Created")

        # Writing Geographic Distribution table
        # current_row = _construct_geographic_distribution(
        #     roles=list(release.team_composition[0].teams[0].model_dump().keys())[1:],
        #     team_list=release.team_composition,
        #     start_row=current_row,
        #     ws=ws,
        #     hours_per_worker=160  # Adjust based on your sprint duration
        # )
        # logger.info("Table created for geographic distribution")
        
        # current_row += 2
    
    try:
        # Convert to Path object for better path handling
        path = Path(output_path)
        # Ensure the file has .xlsx extension
        if path.suffix.lower() != '.xlsx':
            path = path.with_suffix('.xlsx')
        # Validate filename (remove invalid characters)
        invalid_chars = '<>:"/\\|?*'
        clean_name = ''.join(c for c in path.name if c not in invalid_chars)
        path = path.parent / clean_name
        
        # Create parent directories if they don't exist
        path.parent.mkdir(parents=True, exist_ok=True)
        
        # Save the workbook
        logger.info(f"Saving file under path {path}")
        wb.save(path)
        logger.info(f"File saved successfully: {path.absolute()}")
        return True
        
    except PermissionError:
        logger.error(f"Error: Cannot save to {path}. The file may be open in another application.")
        return False
    except OSError as e:
        logger.error(f"Error saving file: {e}")
        return False
    except Exception as e:
        logger.error(f"Unexpected error: {e}")
        return False
    finally:
        # Release resources
        wb.close()
